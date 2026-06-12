"""FastAPI bridge for the React CMF interface.

Streamlit remains available as a backup UI. This API intentionally reuses the
existing repositories and SQLite database so the business logic is not forked.
"""

from __future__ import annotations

import sys
from pathlib import Path
from io import BytesIO
from typing import Any, Dict, List, Optional
import re
import base64
import csv
import hashlib
import secrets
from datetime import date, datetime

APP_DIR = Path(__file__).resolve().parent
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

import db_sqlite

db_sqlite.DB_PATH = APP_DIR / "data" / "cmf.db"

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from repositories.cmf_record_repository_sqlite import CMFRecordRepository
from repositories.project_column_repository_sqlite import ProjectColumnRepository
from repositories.project_repository_sqlite import ProjectRepository
from repositories.user_repository_sqlite import UserRepository
from services.master_schema import STANDARD_COLUMNS_WITH_ROLES, get_column_section

app = FastAPI(title="CMF API", version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost:5174",
        "http://127.0.0.1:5174",
        "http://localhost:5175",
        "http://127.0.0.1:5175",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

project_repo = ProjectRepository()
record_repo = CMFRecordRepository()
column_repo = ProjectColumnRepository()
user_repo = UserRepository()

REACT_CMF_TEMPLATE_PATH = APP_DIR / "data" / "templates" / "CMF_MASTER_TEMPLATE.xlsx"
REACT_DEFAULT_VIEW_COLS = [
    "N° SOURCING, RFQ,ODM,FETE …",
    "APQP GRID",
    "USE CASES",
    "PART NUMBER",
    "WEEKLY CAPACITY CONTRACTED (Parts/Week)",
    "GOR (Green, Orange, Red) Supplier Capacity Contracted regarding Buyer Capacity Requested",
    "CAT1/2/3 VALUATION (G;O;R)",
]
REACT_SECTION_MARKER_COLUMNS = {
    "PART DATA",
    "PART DATA (STEP 1)",
    "SUPPLIER INFORMATION",
    "WEEKLY CONTRACTED CAPACITY",
    "CAPACITY SIZING",
    "CAPACITY WORKSHOP",
    "CAT",
    "CUSTOMIZED COLONNE",
}
GOR_COLUMN = "GOR (Green, Orange, Red) Supplier Capacity Contracted regarding Buyer"
GOR_REQUESTED_COLUMN = "GOR (Green, Orange, Red) Supplier Capacity Contracted regarding Buyer Capacity Requested"
CAT_VALUATION_COLUMN = "CAT1/2/3 VALUATION (G;O;R)"
CONTRACTED_COLUMN = "WEEKLY CAPACITY CONTRACTED (Parts/Week)"
REQUESTED_COLUMN = "LAST WEEKLY CAPACITY REQUESTED"
MEASURED_COLUMN = "WEEKLY CAPACITY MEASURED"
DASHBOARD_REQUIRED_FIELDS = [
    "SUPPLIER NAME",
    "PART NUMBER",
    "PROGRAM BUYER / BUYER",
    "SQE",
    CONTRACTED_COLUMN,
    REQUESTED_COLUMN,
    GOR_REQUESTED_COLUMN,
    "CAT FORECASTED DATE",
    "SHARED FOLDER",
]


class CustomColumnPayload(BaseModel):
    column_name: str
    owner_role: str
    section: Optional[str] = None
    actor_email: Optional[str] = None


class ProjectCreate(BaseModel):
    project: str
    part_of_project: str
    capacity_manager_name: str
    buyer_assigned_name: Optional[str] = None
    sqd_assigned_name: Optional[str] = None
    description: Optional[str] = None
    cmf_status: str = "ACTIVE"
    created_by: Optional[str] = None
    custom_columns: List[CustomColumnPayload] = []


class RecordUpsert(BaseModel):
    part_number: str
    apqp_grid: Optional[str] = None
    values: Dict[str, Any] = {}
    updated_by: str = "react-ui"
    actor_email: Optional[str] = None


class AdminRecordDirectSave(BaseModel):
    part_number: str
    apqp_grid: Optional[str] = None
    values: Dict[str, Any] = {}
    updated_by: str = "admin"


class ProjectRecordDirectSave(BaseModel):
    part_number: Optional[str] = None
    apqp_grid: Optional[str] = None
    values: Dict[str, Any] = {}
    updated_by: str = "react-ui"
    actor_email: Optional[str] = None


class ActionPlanCreate(BaseModel):
    record_id: int
    problem_type: str
    action_text: str
    old_value: Optional[str] = None
    new_value: str
    created_by: str = "react-ui"
    actor_email: Optional[str] = None


class ValueUpdate(BaseModel):
    values: Dict[str, Any]
    updated_by: str = "react-ui"


class ProjectUpdate(BaseModel):
    name: Optional[str] = None
    capacity_manager_name: Optional[str] = None
    buyer_assigned_name: Optional[str] = None
    sqd_assigned_name: Optional[str] = None
    supplier_name: Optional[str] = None
    cmf_status: Optional[str] = None
    description: Optional[str] = None
    actor_email: Optional[str] = None


class UserCreate(BaseModel):
    email: str
    full_name: Optional[str] = None
    role: str = "BUYER"
    password: str


class UserRoleUpdate(BaseModel):
    role: str


class UserPasswordUpdate(BaseModel):
    password: str


class LoginPayload(BaseModel):
    email: str
    password: str


class RoleRecordSave(BaseModel):
    part_number: str
    apqp_grid: Optional[str] = None
    values: Dict[str, Any]
    role: str
    section: Optional[str] = None
    updated_by: str = "react-ui"
    actor_email: Optional[str] = None
    create_if_missing: bool = True


class ImportFileParse(BaseModel):
    filename: str
    content_base64: str


def _project_code(project: str, part_of_project: str) -> str:
    return f"{project.strip()} - {part_of_project.strip()}"


def _split_project_code(code: str, name: str) -> Dict[str, str]:
    suffix = f" - {name}"
    if code.endswith(suffix):
        return {"project": code[: -len(suffix)], "part_of_project": name}
    if " - " in code:
        project, part = code.split(" - ", 1)
        return {"project": project, "part_of_project": part}
    return {"project": code, "part_of_project": name}


def _record_payload(record) -> Dict[str, Any]:
    data = record.to_dict()
    values = {key: value for key, value in data.items() if key not in {"id", "project_id", "apqp_grid", "apqp", "part_number", "status", "created_at", "updated_at", "updated_by"}}
    return {
        "id": record.id,
        "project_id": record.project_id,
        "apqp_grid": record.apqp_grid,
        "part_number": record.part_number,
        "status": record.status,
        "created_at": record.created_at,
        "updated_at": record.updated_at,
        "updated_by": record.updated_by,
        "values": values,
        "flat": data,
    }


def _project_payload(project) -> Dict[str, Any]:
    names = _split_project_code(project.code, project.name)
    records = record_repo.get_records_for_project(project.id)
    completion_values: List[float] = []
    for record in records:
        flat = record.to_dict()
        filled = len([value for value in flat.values() if value not in (None, "")])
        completion_values.append(min(100, (filled / 18) * 100))

    completion = round(sum(completion_values) / len(completion_values)) if completion_values else 0
    return {
        "id": project.id,
        "code": project.code,
        "name": project.name,
        **names,
        "description": project.description,
        "capacity_manager_name": project.capacity_manager_name,
        "buyer_assigned_name": project.buyer_assigned_name,
        "sqd_assigned_name": project.sqd_assigned_name,
        "supplier_name": project.supplier_name,
        "cmf_status": project.cmf_status,
        "created_by": project.created_by,
        "created_at": project.created_at,
        "updated_at": project.updated_at,
        "records_count": len(records),
        "completion": completion,
    }


def _user_payload(user) -> Dict[str, Any]:
    return {
        "id": user.id,
        "email": user.email,
        "full_name": user.full_name,
        "role": user.role,
        "created_at": user.created_at,
    }


def _actor_user(actor_email: Optional[str]) -> Optional[Dict[str, Any]]:
    if not actor_email:
        return None
    conn = db_sqlite.get_connection()
    try:
        row = conn.execute(
            "SELECT id, email, full_name, role FROM app_users WHERE lower(email) = lower(?)",
            (actor_email.strip(),),
        ).fetchone()
        return dict(row) if row else None
    finally:
        db_sqlite.close_connection(conn)


def _same_assignment(actor: Dict[str, Any], assigned_name: Optional[str]) -> bool:
    assigned = (assigned_name or "").strip().lower()
    if not assigned or assigned.startswith("unassigned"):
        return False
    email = (actor.get("email") or "").strip().lower()
    full_name = (actor.get("full_name") or "").strip().lower()
    return assigned == email or assigned == full_name or email in assigned or bool(full_name and full_name in assigned)


def _ensure_project_write_access(project, actor_email: Optional[str], expected_role: Optional[str] = None) -> Dict[str, Any]:
    actor = _actor_user(actor_email)
    if not actor:
        raise HTTPException(status_code=403, detail="User is required to modify this project")

    actor_role = (actor.get("role") or "").upper()
    if actor_role in {"ADMIN", "SUPER_ADMIN"}:
        return actor
    if expected_role and actor_role != expected_role.upper():
        raise HTTPException(status_code=403, detail="Your role cannot modify this project section")

    allowed = False
    if expected_role == "BUYER":
        allowed = _same_assignment(actor, project.buyer_assigned_name)
    elif expected_role == "SQD":
        allowed = _same_assignment(actor, project.sqd_assigned_name)
    elif expected_role == "CAPACITY_MANAGER":
        allowed = _same_assignment(actor, project.capacity_manager_name)
    else:
        allowed = (
            _same_assignment(actor, project.capacity_manager_name)
            or _same_assignment(actor, project.buyer_assigned_name)
            or _same_assignment(actor, project.sqd_assigned_name)
        )

    if not allowed:
        raise HTTPException(status_code=403, detail="Read-only: user is not assigned to this project")
    return actor


def _hash_password(password: str, salt: Optional[str] = None) -> str:
    salt_value = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt_value.encode("utf-8"), 120_000)
    return f"pbkdf2_sha256${salt_value}${digest.hex()}"


def _verify_password(password: str, stored_hash: str) -> bool:
    if not stored_hash:
        return False
    if stored_hash.startswith("pbkdf2_sha256$"):
        try:
            _, salt, digest = stored_hash.split("$", 2)
        except ValueError:
            return False
        return secrets.compare_digest(_hash_password(password, salt), stored_hash)
    # Compatibility for old rows that stored a raw temporary value.
    return secrets.compare_digest(password, stored_hash)


def _ensure_users_table() -> None:
    conn = db_sqlite.get_connection()
    try:
        with conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS app_users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    email TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    full_name TEXT,
                    role TEXT NOT NULL DEFAULT 'BUYER',
                    created_at TEXT DEFAULT (datetime('now'))
                )
                """
            )
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) AS count FROM app_users")
            count = int(cur.fetchone()["count"] or 0)
            if count == 0:
                conn.execute(
                    """
                    INSERT INTO app_users (email, password_hash, full_name, role)
                    VALUES (?, ?, ?, ?)
                    """,
                    ("admin@cmf.local", _hash_password("admin123"), "CMF Administrator", "ADMIN"),
                )
    finally:
        db_sqlite.close_connection(conn)


def _ensure_react_schema_migrations() -> None:
    conn = db_sqlite.get_connection()
    try:
        with conn:
            conn.execute(
                """
                UPDATE project_columns
                SET column_name = ?
                WHERE column_name = ?
                  AND NOT EXISTS (
                      SELECT 1
                      FROM project_columns AS existing
                      WHERE existing.project_id = project_columns.project_id
                        AND existing.column_name = ?
                  )
                """,
                ("CAPACITY SOURCE", "PEAK YEAR", "CAPACITY SOURCE"),
            )
            conn.execute(
                """
                DELETE FROM project_columns
                WHERE column_name = ?
                  AND EXISTS (
                      SELECT 1
                      FROM project_columns AS existing
                      WHERE existing.project_id = project_columns.project_id
                        AND existing.column_name = ?
                  )
                """,
                ("PEAK YEAR", "CAPACITY SOURCE"),
            )
            conn.execute(
                """
                UPDATE cmf_record_values
                SET column_name = ?
                WHERE column_name = ?
                  AND NOT EXISTS (
                      SELECT 1
                      FROM cmf_record_values AS existing
                      WHERE existing.record_id = cmf_record_values.record_id
                        AND existing.column_name = ?
                  )
                """,
                ("CAPACITY SOURCE", "PEAK YEAR", "CAPACITY SOURCE"),
            )
            conn.execute(
                """
                DELETE FROM cmf_record_values
                WHERE column_name = ?
                  AND EXISTS (
                      SELECT 1
                      FROM cmf_record_values AS existing
                      WHERE existing.record_id = cmf_record_values.record_id
                        AND existing.column_name = ?
                  )
                """,
                ("PEAK YEAR", "CAPACITY SOURCE"),
            )
            projects = conn.execute("SELECT id FROM projects").fetchall()
            column_names = {row["name"] for row in conn.execute("PRAGMA table_info(project_columns)").fetchall()}
            if "custom_section" not in column_names:
                conn.execute("ALTER TABLE project_columns ADD COLUMN custom_section TEXT")
            for col_def in STANDARD_COLUMNS_WITH_ROLES:
                conn.execute(
                    "UPDATE project_columns SET custom_section = ? WHERE column_name = ? AND (custom_section IS NULL OR custom_section = '')",
                    (col_def["section"], col_def["name"]),
                )
            for project in projects:
                project_id = project["id"]
                pn_row = conn.execute(
                    "SELECT display_order FROM project_columns WHERE project_id = ? AND column_name = ?",
                    (project_id, "PART NUMBER"),
                ).fetchone()
                part_name_order = int(pn_row["display_order"] or 0) + 1 if pn_row else None
                exists = conn.execute(
                    "SELECT display_order FROM project_columns WHERE project_id = ? AND column_name = ?",
                    (project_id, "PART NAME"),
                ).fetchone()
                if part_name_order is not None and (not exists or int(exists["display_order"] or -1) != part_name_order):
                    conn.execute(
                        """
                        UPDATE project_columns
                        SET display_order = display_order + 1
                        WHERE project_id = ?
                          AND display_order >= ?
                          AND column_name != ?
                        """,
                        (project_id, part_name_order, "PART NAME"),
                    )
                if not exists:
                    order_row = conn.execute(
                        "SELECT COALESCE(MAX(display_order), -1) + 1 AS next_order FROM project_columns WHERE project_id = ?",
                        (project_id,),
                    ).fetchone()
                    display_order = part_name_order if part_name_order is not None else int(order_row["next_order"] or 0)
                    conn.execute(
                        """
                        INSERT OR IGNORE INTO project_columns (
                            project_id, column_name, owner_role, is_auto, display_order
                        ) VALUES (?, ?, ?, ?, ?)
                        """,
                        (project_id, "PART NAME", "BUYER", 0, display_order),
                    )
                elif part_name_order is not None:
                    conn.execute(
                        """
                        UPDATE project_columns
                        SET display_order = ?
                        WHERE project_id = ? AND column_name = ?
                        """,
                        (part_name_order, project_id, "PART NAME"),
                    )
                conn.execute(
                    """
                    INSERT OR IGNORE INTO project_column_permissions (
                        project_id, column_name, role, can_edit
                    ) VALUES (?, ?, ?, ?)
                    """,
                    (project_id, "PART NAME", "BUYER", 1),
                )
    finally:
        db_sqlite.close_connection(conn)


def _write_audit_log(
    action: str,
    entity_type: str,
    entity_id: Optional[int] = None,
    user_name: str = "admin",
    old_value: Optional[str] = None,
    new_value: Optional[str] = None,
    project_id: Optional[int] = None,
) -> None:
    conn = db_sqlite.get_connection()
    try:
        with conn:
            conn.execute(
                """
                INSERT INTO audit_logs (
                    action, entity_type, entity_id, user_name, old_value, new_value, project_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (action, entity_type, entity_id, user_name, old_value, new_value, project_id),
            )
    finally:
        db_sqlite.close_connection(conn)


def _audit_rows(limit: int = 250) -> List[Dict[str, Any]]:
    conn = db_sqlite.get_connection()
    try:
        rows = conn.execute(
            """
            SELECT id, action, entity_type, entity_id, user_name, old_value, new_value, timestamp, project_id
            FROM audit_logs
            ORDER BY id DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
        return [dict(row) for row in rows]
    finally:
        db_sqlite.close_connection(conn)


def _action_plan_rows(project_id: int) -> List[Dict[str, Any]]:
    conn = db_sqlite.get_connection()
    try:
        rows = conn.execute(
            """
            SELECT id, project_id, record_id, part_number, problem_type, target_column,
                   action_text, old_value, new_value, created_by, created_at
            FROM action_plans
            WHERE project_id = ?
            ORDER BY id DESC
            """,
            (project_id,),
        ).fetchall()
        return [dict(row) for row in rows]
    finally:
        db_sqlite.close_connection(conn)


def _safe_excel_value(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", str(value))


def _normalized_header(value: Any) -> str:
    return re.sub(r"\s+", " ", _safe_excel_value(value)).strip()


def _first_number(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    match = re.search(r"-?\d+(?:[.,]\d+)?", str(value))
    if not match:
        return None
    try:
        return float(match.group(0).replace(",", "."))
    except ValueError:
        return None


def _react_gor_value(row: Dict[str, Any]) -> str:
    contracted = _first_number(row.get(CONTRACTED_COLUMN))
    requested = _first_number(row.get(REQUESTED_COLUMN))
    if contracted is None or requested is None or requested <= 0:
        return ""
    existing = row.get(GOR_COLUMN)
    if existing not in (None, ""):
        return str(existing)
    ratio = contracted / requested
    if ratio >= 1:
        return "G"
    if ratio >= 0.8:
        return "O"
    return "R"


def _react_cat_value(row: Dict[str, Any]) -> str:
    contracted = _first_number(row.get(CONTRACTED_COLUMN))
    requested = _first_number(row.get(REQUESTED_COLUMN))
    measured = _first_number(row.get(MEASURED_COLUMN))
    if contracted is None or requested is None or measured is None or requested <= 0:
        return ""
    existing = row.get(CAT_VALUATION_COLUMN)
    if existing not in (None, ""):
        return str(existing)
    tolerance = max(abs(requested) * 0.05, 1e-9)
    if abs(measured - requested) <= tolerance:
        return "O"
    if measured > requested:
        return "G"
    return "R"


def _template_aliases(header: str) -> List[str]:
    normalized = _normalized_header(header)
    aliases = {
        "PART DATA": ["PART DATA "],
        "PART NAME": ["part_name"],
        GOR_REQUESTED_COLUMN: [GOR_COLUMN],
        "SHARED FOLDER": ["SCR - SHARED FOLDER"],
        "SHARED FOLDER DATE (DD/MM/YYYY)": ["SCR DATE (DD/MM/YYYY)"],
        "SCR DATE (DD/MM/YYYY)": ["SCR DATE (DD/MM/YYYY)", "SHARED FOLDER DATE (DD/MM/YYYY)"],
        "CAPACITY SOURCE": ["PEAK YEAR"],
        "CMF LINE N°": ["CMF LINE NÂ°"],
        "N° SOURCING, RFQ,ODM,FETE …": ["NÂ° SOURCING, RFQ,ODM,FETE â€¦"],
        "PROGRAM BUYER / BUYER": ["PROGRAM BUYER / BUYER"],
        "WEEKLY CAPACITY CONTRACTED (Parts/Week)": ["WEEKLY CAPACITY CONTRACTED (Parts/Week)"],
        "CAPACITY CONTRACTED STEP (Parts/Week)": ["CONTRACTED CAPACITY STEP (Parts/Week)", "CAPACITY CONTRACTED STEP (Parts/Week)"],
        "GOR (Green, Orange, Red) Supplier Capacity Contracted regarding Buyer Capacity Requested": [
            "GOR (Green, Orange, Red) Supplier Capacity Contracted regarding Buyer"
        ],
        "SCR DATE (DD/MM/YYYY)": ["SCR DATE (DD/MM/YYYY)"],
        "SUPERMIX ACTIVITY OUI/NON": ["SUPERMIX ACTIVITY OUI/NON"],
        "SUPERMIX ACTIVITY Weekly Capacity": ["SUPERMIX ACTIVITY Weekly Capacity"],
        "SUPERMIX ACTIVITY Commentaires": ["SUPERMIX ACTIVITY Commentaires"],
        "CAPACITY WORKSHOP Done date": ["CAPACITY WORKSHOP Done date"],
        "APQP Grid Project (Project Wave x)": ["APQP Grid Project (Project Wave x)"],
        "CAT1 FORECASTED DATE (YYCWxx)": ["CAT1 FORECASTED DATE (YYCWxx)"],
        "CAT2 FORECASTED DATE (YYCWxx)": ["CAT2 FORECASTED DATE (YYCWxx)"],
        "CAT3 FORECASTED DATE (YYCWxx)": ["CAT3 FORECASTED DATE (YYCWxx)"],
        "CAT1 REALISED DATE (YYCWxx)": ["CAT1 REALISED DATE (YYCWxx)"],
        "CAT2 REALISED DATE (YYCWxx)": ["CAT2 REALISED DATE (YYCWxx)"],
        "CAT3 REALISED DATE (YYCWxx)": ["CAT3 REALISED DATE (YYCWxx)"],
    }
    aliases.update(
        {
            "N° SOURCING, RFQ,ODM,FETE …": [
                "NÂ° SOURCING, RFQ,ODM,FETE â€¦",
                "NÃ‚Â° SOURCING, RFQ,ODM,FETE Ã¢â‚¬Â¦",
            ],
            "APQP GRID": ["apqp_grid"],
            "PART NUMBER": ["part_number"],
        }
    )
    values = [normalized, *aliases.get(normalized, [])]
    unique = []
    for value in values:
        if value and value not in unique:
            unique.append(value)
    return unique


def _project_label(project: Dict[str, Any]) -> str:
    code = str(project.get("code") or "").strip()
    name = str(project.get("name") or "").strip()
    if code and name and code.endswith(f" - {name}"):
        return code
    if code and name:
        return f"{code} - {name}"
    return code or name


def _roadmap_labels(limit: int = 3) -> List[str]:
    return [_project_label(project) for project in record_repo.get_all_projects_for_cross_view()[:limit]]


def _project_custom_column_names(project_id: int) -> List[str]:
    names = []
    for column in column_repo.get_columns(project_id):
        column_name = column["column_name"]
        if column_name not in {col["name"] for col in STANDARD_COLUMNS_WITH_ROLES} and column_name not in names:
            names.append(column_name)
    return names


def _column_section_for_project(project_id: int, column_name: str) -> str:
    standard_section = get_column_section(column_name)
    if standard_section != "CUSTOMIZED COLUMNS":
        return standard_section
    for column in column_repo.get_columns(project_id):
        if column["column_name"] == column_name:
            return column.get("custom_section") or column.get("section") or "CUSTOMIZED COLUMNS"
    return "CUSTOMIZED COLUMNS"


def _is_custom_placeholder(column_name: str) -> bool:
    return bool(re.fullmatch(r"Colonne\d+", _normalized_header(column_name), flags=re.IGNORECASE))


def _display_template_header(column_name: str, project_id: Optional[int] = None) -> str:
    normalized = _normalized_header(column_name)
    roadmap_idx = {
        "Project - Part of Project": 0,
        "Project - Part of Project1": 0,
        "Project - Part of Project2": 1,
        "Project - Part of Project3": 2,
    }.get(normalized)
    labels = _roadmap_labels()
    if roadmap_idx is not None and roadmap_idx < len(labels):
        return labels[roadmap_idx]
    return normalized


def _add_custom_column_to_project(project_id: int, column_name: str, owner_role: str) -> Dict[str, Any]:
    return _add_custom_column_to_project_section(project_id, column_name, owner_role, None)


def _add_custom_column_to_project_section(project_id: int, column_name: str, owner_role: str, section: Optional[str]) -> Dict[str, Any]:
    normalized_name = str(column_name or "").strip()
    normalized_role = str(owner_role or "").strip().upper()
    allowed_roles = {"BUYER", "SQD", "CAPACITY_MANAGER", "ADMIN"}
    role_sections = {
        "BUYER": {"PART DATA", "WEEKLY CONTRACTED CAPACITY", "CAPACITY SIZING"},
        "CAPACITY_MANAGER": {"CAPACITY SIZING", "CAPACITY WORKSHOP (STEP 2)"},
        "SQD": {"PART DATA", "SUPPLIER INFORMATION", "CAPACITY WORKSHOP (STEP 2)", "CAT"},
        "ADMIN": {"PART DATA", "WEEKLY CONTRACTED CAPACITY", "CAPACITY SIZING", "CAPACITY WORKSHOP (STEP 2)", "SUPPLIER INFORMATION", "CAT"},
    }
    if not normalized_name:
        raise HTTPException(status_code=400, detail="Column name is required")
    if normalized_role not in allowed_roles:
        raise HTTPException(status_code=400, detail="owner_role must be BUYER, SQD, CAPACITY_MANAGER, or ADMIN")
    normalized_section = str(section or "").strip().upper()
    if not normalized_section:
        normalized_section = sorted(role_sections[normalized_role])[0]
    valid_sections = role_sections[normalized_role]
    if normalized_section not in valid_sections:
        raise HTTPException(status_code=400, detail=f"section must be one of: {', '.join(sorted(valid_sections))}")

    existing_columns = column_repo.get_columns(project_id)
    next_order = max([int(column.get("display_order") or 0) for column in existing_columns], default=-1) + 1
    inserted = column_repo.insert(
        project_id=project_id,
        column_name=normalized_name,
        owner_roles=[normalized_role],
        is_auto=0,
        display_order=next_order,
        can_edit=1,
        custom_section=normalized_section,
    )
    if not inserted:
        raise HTTPException(status_code=400, detail="Unable to add custom column")
    column = inserted[0]
    return {
        "id": column.id,
        "project_id": column.project_id,
        "column_name": column.column_name,
        "owner_role": column.owner_role,
        "is_auto": column.is_auto,
        "display_order": column.display_order,
        "section": normalized_section,
        "roles": column_repo.get_roles_for_column(project_id, column.column_name),
    }


def _roadmap_context() -> Dict[str, Any]:
    projects = record_repo.get_all_projects_for_cross_view()
    cross_data = record_repo.get_cross_project_part_number_view()
    by_part_number = {
        str(entry.get("part_number") or "").strip(): entry
        for entry in cross_data
        if str(entry.get("part_number") or "").strip()
    }
    return {"projects": projects, "by_part_number": by_part_number, "cross_data": cross_data}


def _apply_roadmap(flat: Dict[str, Any], context: Dict[str, Any]) -> Dict[str, Any]:
    enriched = dict(flat)
    part_number = str(enriched.get("PART NUMBER") or enriched.get("part_number") or "").strip()
    enriched["PART NUMBER"] = part_number
    enriched["PART NAME"] = enriched.get("PART NAME") or enriched.get("part_name") or ""
    enriched["APQP GRID"] = enriched.get("APQP GRID") or enriched.get("apqp_grid") or enriched.get("apqp") or ""
    gor_value = _react_gor_value(enriched)
    cat_value = _react_cat_value(enriched)
    enriched[GOR_COLUMN] = gor_value
    enriched[GOR_REQUESTED_COLUMN] = gor_value
    enriched[CAT_VALUATION_COLUMN] = cat_value
    entry = context["by_part_number"].get(part_number, {})
    projects = context["projects"]
    enriched["ROADMAP"] = ""
    for idx, project in enumerate(projects[:3], 1):
        template_col = "Project - Part of Project" if idx == 1 else f"Project - Part of Project{idx}"
        placeholder_col = f"Project Code - Project Name {idx}"
        project_label = _project_label(project) or template_col
        value = "X" if entry.get(f"proj_{project['id']}", False) else ""
        enriched[template_col] = value
        enriched[placeholder_col] = value
        enriched[project_label] = value
    count = int(entry.get("project_count") or sum(1 for key, value in entry.items() if key.startswith("proj_") and value))
    enriched["CarryOver - Adapted"] = "Adapted" if count == 1 else ("CarryOver" if count > 1 else "")
    return enriched


def _template_columns() -> List[str]:
    if REACT_CMF_TEMPLATE_PATH.exists():
        try:
            from openpyxl import load_workbook

            wb = load_workbook(REACT_CMF_TEMPLATE_PATH, read_only=True, data_only=False)
            ws = wb["CMF"] if "CMF" in wb.sheetnames else wb.active
            header_row = 1 if any(_normalized_header(ws.cell(row=2, column=col_idx).value) == "xxx" for col_idx in range(1, ws.max_column + 1)) else 3
            columns = []
            for col_idx in range(1, ws.max_column + 1):
                column = _normalized_header(ws.cell(row=header_row, column=col_idx).value)
                marker = _normalized_header(ws.cell(row=header_row + 1, column=col_idx).value) if header_row == 1 else ""
                if not column or marker.lower() == "xxx" or column in REACT_SECTION_MARKER_COLUMNS:
                    continue
                if _is_custom_placeholder(column):
                    continue
                columns.append(column)
            template_columns = [_display_template_header(column) for column in columns]
            custom_columns = [
                column["column_name"]
                for project in project_repo.get_all_projects()
                for column in column_repo.get_columns(project.id)
                if column["column_name"] not in {col["name"] for col in STANDARD_COLUMNS_WITH_ROLES}
                and column["column_name"] not in template_columns
            ]
            return template_columns + list(dict.fromkeys(custom_columns))
        except Exception:
            pass
    return [col["name"] for col in STANDARD_COLUMNS_WITH_ROLES if col["name"] not in REACT_SECTION_MARKER_COLUMNS]


def _project_full_rows(project_id: int) -> List[Dict[str, Any]]:
    context = _roadmap_context()
    rows = []
    for index, record in enumerate(record_repo.get_records_for_project(project_id), 1):
        row = _apply_roadmap(record.to_dict(), context)
        row["__record_id"] = record.id
        row["CMF LINE NÂ°"] = index
        row["CMF LINE N°"] = index
        rows.append(row)
    return rows


def _display_rows(rows: List[Dict[str, Any]], columns: List[str]) -> List[Dict[str, Any]]:
    return [
        {"__record_id": row.get("__record_id"), **{column: _template_cell_value(row, column) for column in columns}}
        for row in rows
    ]


def _dash_value(row: Dict[str, Any], *headers: str) -> str:
    for header in headers:
        value = _template_cell_value(row, header)
        if str(value).strip():
            return str(value).strip()
    return ""


def _dash_number(row: Dict[str, Any], *headers: str) -> Optional[float]:
    for header in headers:
        value = _first_number(_dash_value(row, header))
        if value is not None:
            return value
    return None


def _normalize_risk(value: Any) -> str:
    text_value = str(value or "").strip().upper()
    if text_value in {"G", "GREEN", "VERT", "VERTE"}:
        return "Green"
    if text_value in {"O", "ORANGE"}:
        return "Orange"
    if text_value in {"R", "RED", "ROUGE"}:
        return "Red"
    return "Unknown"


def _capacity_status(contracted: Optional[float], requested: Optional[float]) -> str:
    if contracted is None or requested is None or requested <= 0:
        return "Unknown"
    coverage = contracted / requested
    if coverage >= 1:
        return "Green"
    if coverage >= 0.9:
        return "Orange"
    return "Red"


def _parse_dashboard_date(value: Any) -> Optional[date]:
    text_value = str(value or "").strip()
    if not text_value:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text_value, fmt).date()
        except ValueError:
            pass
    match = re.search(r"(?P<year>\d{2,4})\s*(?:CW|W)\s*(?P<week>\d{1,2})", text_value, re.IGNORECASE)
    if match:
        year = int(match.group("year"))
        if year < 100:
            year += 2000
        week = max(1, min(53, int(match.group("week"))))
        try:
            return date.fromisocalendar(year, week, 1)
        except ValueError:
            return None
    return None


def _cat_single_status(forecast: Any, realised: Any, today: Optional[date] = None) -> str:
    if str(realised or "").strip():
        return "Done"
    forecast_date = _parse_dashboard_date(forecast)
    if not forecast_date:
        return "Not planned"
    today_value = today or date.today()
    return "Late" if forecast_date < today_value else "Planned"


def _dashboard_record(row: Dict[str, Any]) -> Dict[str, Any]:
    requested = _dash_number(row, REQUESTED_COLUMN)
    contracted = _dash_number(row, CONTRACTED_COLUMN, "CAPACITY CONTRACTED STEP (Parts/Week)")
    measured = _dash_number(row, MEASURED_COLUMN)
    gap = (contracted or 0) - (requested or 0) if contracted is not None and requested is not None else None
    coverage = contracted / requested if contracted is not None and requested not in (None, 0) else None
    measured_gap = (measured or 0) - (requested or 0) if measured is not None and requested is not None else None
    measured_coverage = measured / requested if measured is not None and requested not in (None, 0) else None
    calculated_status = _capacity_status(contracted, requested)
    measured_status = _capacity_status(measured, requested)
    gor = _normalize_risk(_dash_value(row, GOR_REQUESTED_COLUMN, GOR_COLUMN)) 
    if gor == "Unknown":
        gor = calculated_status

    cat_statuses = {
        "CAT1": _cat_single_status(_dash_value(row, "CAT1 FORECASTED DATE (YYCWxx)", "CAT1 FORECASTED DATE"), _dash_value(row, "CAT1 REALISED DATE (YYCWxx)", "CAT1 REALISED DATE")),
        "CAT2": _cat_single_status(_dash_value(row, "CAT2 FORECASTED DATE (YYCWxx)", "CAT2 FORECASTED DATE"), _dash_value(row, "CAT2 REALISED DATE (YYCWxx)", "CAT2 REALISED DATE")),
        "CAT3": _cat_single_status(_dash_value(row, "CAT3 FORECASTED DATE (YYCWxx)", "CAT3 FORECASTED DATE"), _dash_value(row, "CAT3 REALISED DATE (YYCWxx)", "CAT3 REALISED DATE")),
    }
    cat_rollup = "Late" if "Late" in cat_statuses.values() else ("Done" if all(status == "Done" for status in cat_statuses.values()) else ("Planned" if "Planned" in cat_statuses.values() else "Not planned"))
    cat_evaluation_status = _normalize_risk(_dash_value(row, CAT_VALUATION_COLUMN, "CAT1/2/3 VALUATION"))

    required_values = [
        _dash_value(row, "SUPPLIER NAME"),
        _dash_value(row, "PART NUMBER"),
        _dash_value(row, "PROGRAM BUYER / BUYER"),
        _dash_value(row, "SQE"),
        _dash_value(row, CONTRACTED_COLUMN),
        _dash_value(row, REQUESTED_COLUMN),
        _dash_value(row, GOR_REQUESTED_COLUMN, GOR_COLUMN),
        _dash_value(row, "CAT1 FORECASTED DATE (YYCWxx)", "CAT2 FORECASTED DATE (YYCWxx)", "CAT3 FORECASTED DATE (YYCWxx)"),
        _dash_value(row, "SHARED FOLDER", "SCR - SHARED FOLDER", "SHARED FOLDER - link"),
    ]
    completeness = round(sum(1 for value in required_values if str(value).strip()) / len(required_values) * 100, 1)

    priority = 4
    if (
        gor == "Red"
        or cat_evaluation_status == "Red"
        or cat_rollup == "Late"
        or (gap is not None and gap < 0 and (coverage is None or coverage < 0.9))
        or (measured_gap is not None and measured_gap < 0 and (measured_coverage is None or measured_coverage < 0.9))
    ):
        priority = 1
    elif gor == "Orange" or cat_evaluation_status == "Orange" or (coverage is not None and 0.9 <= coverage < 1) or (measured_coverage is not None and 0.9 <= measured_coverage < 1):
        priority = 2
    elif gor == "Green" and cat_evaluation_status == "Green" and (completeness < 100 or cat_rollup == "Not planned"):
        priority = 3

    return {
        "record_id": row.get("__record_id"),
        "supplier_name": _dash_value(row, "SUPPLIER NAME"),
        "country": _dash_value(row, "COUNTRY"),
        "location": _dash_value(row, "LOCATION"),
        "part_number": _dash_value(row, "PART NUMBER"),
        "part_name": _dash_value(row, "PART NAME"),
        "buyer": _dash_value(row, "PROGRAM BUYER / BUYER"),
        "sqe": _dash_value(row, "SQE"),
        "apqp_grid_project": _dash_value(row, "APQP Grid Project (Project Wave x)", "APQP Grid Project"),
        "use_cases": _dash_value(row, "USE CASES"),
        "year_of_max_need": _dash_value(row, "YEAR OF MAX NEED"),
        "capacity_source": _dash_value(row, "CAPACITY SOURCE"),
        "last_weekly_capacity_requested": requested,
        "weekly_capacity_contracted": contracted,
        "weekly_capacity_measured": measured,
        "capacity_gap": gap,
        "coverage_rate": coverage,
        "capacity_status": calculated_status,
        "measured_capacity_gap": measured_gap,
        "measured_coverage_rate": measured_coverage,
        "measured_capacity_status": measured_status,
        "gor": gor,
        "cat_status": cat_rollup,
        "cat_statuses": cat_statuses,
        "cat_valuation": cat_evaluation_status,
        "cat_evaluation_status": cat_evaluation_status,
        "capacity_workshop_done_date": _dash_value(row, "CAPACITY WORKSHOP Done date"),
        "shared_folder_link": _dash_value(row, "SHARED FOLDER - link", "SHARED FOLDER", "SCR - SHARED FOLDER"),
        "comments": _dash_value(row, "Comments"),
        "data_completeness": completeness,
        "priority": priority,
    }


def _dashboard_filtered_records(project_id: int, filters: Dict[str, Optional[str]]) -> List[Dict[str, Any]]:
    records = [_dashboard_record(row) for row in _project_full_rows(project_id)]
    filter_fields = {
        "supplier": "supplier_name",
        "part_number": "part_number",
        "country": "country",
        "location": "location",
        "buyer": "buyer",
        "sqe": "sqe",
        "gor": "gor",
        "apqp": "apqp_grid_project",
        "use_cases": "use_cases",
        "year": "year_of_max_need",
        "cat_status": "cat_status",
        "cat_evaluation_status": "cat_evaluation_status",
        "capacity_source": "capacity_source",
    }
    for query_key, record_key in filter_fields.items():
        expected = (filters.get(query_key) or "").strip().lower()
        if expected:
            records = [record for record in records if str(record.get(record_key) or "").strip().lower() == expected]
    return records


def _count_by(records: List[Dict[str, Any]], key: str) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for record in records:
        value = str(record.get(key) or "Unknown").strip() or "Unknown"
        counts[value] = counts.get(value, 0) + 1
    return counts


def _top_counts(records: List[Dict[str, Any]], key: str, risk: str = "Red", limit: int = 10) -> List[Dict[str, Any]]:
    counts = _count_by([record for record in records if record.get("gor") == risk], key)
    return [{"label": label, "value": value} for label, value in sorted(counts.items(), key=lambda item: item[1], reverse=True)[:limit]]


def _stacked_risk_by(records: List[Dict[str, Any]], group_key: str, status_key: str = "cat_evaluation_status", limit: int = 10) -> List[Dict[str, Any]]:
    grouped: Dict[str, Dict[str, int]] = {}
    for record in records:
        label = str(record.get(group_key) or "Unknown").strip() or "Unknown"
        status = str(record.get(status_key) or "Unknown").strip() or "Unknown"
        grouped.setdefault(label, {"Green": 0, "Orange": 0, "Red": 0, "Unknown": 0})
        grouped[label][status if status in grouped[label] else "Unknown"] += 1
    ranked = sorted(grouped.items(), key=lambda item: (item[1]["Red"], item[1]["Orange"], item[1]["Green"]), reverse=True)
    return [{"label": label, **counts} for label, counts in ranked[:limit]]


def _dashboard_payload(project_id: int, filters: Dict[str, Optional[str]]) -> Dict[str, Any]:
    records = _dashboard_filtered_records(project_id, filters)
    all_records = [_dashboard_record(row) for row in _project_full_rows(project_id)]
    total_requested = sum(record["last_weekly_capacity_requested"] or 0 for record in records)
    total_contracted = sum(record["weekly_capacity_contracted"] or 0 for record in records)
    total_measured = sum(record["weekly_capacity_measured"] or 0 for record in records)
    missing_capacity = sum(record["capacity_gap"] or 0 for record in records if (record["capacity_gap"] or 0) < 0)
    total_measured_gap = sum(record["measured_capacity_gap"] or 0 for record in records)
    missing_measured_capacity = sum(record["measured_capacity_gap"] or 0 for record in records if (record["measured_capacity_gap"] or 0) < 0)
    gor_distribution = {status: 0 for status in ["Green", "Orange", "Red", "Unknown"]}
    gor_distribution.update(_count_by(records, "gor"))
    measured_status_distribution = {status: 0 for status in ["Green", "Orange", "Red", "Unknown"]}
    measured_status_distribution.update(_count_by(records, "measured_capacity_status"))
    cat_evaluation_distribution = {status: 0 for status in ["Green", "Orange", "Red", "Unknown"]}
    cat_evaluation_distribution.update(_count_by(records, "cat_evaluation_status"))

    cat_status = {
        cat: {status: 0 for status in ["Done", "Late", "Planned", "Not planned"]}
        for cat in ["CAT1", "CAT2", "CAT3"]
    }
    for record in records:
        for cat, status in record["cat_statuses"].items():
            cat_status[cat][status] += 1

    top_gaps = sorted(
        [
            {
                "label": record["supplier_name"] or record["part_number"] or "Unknown",
                "part_number": record["part_number"],
                "value": record["capacity_gap"] or 0,
            }
            for record in records
            if (record["capacity_gap"] or 0) < 0
        ],
        key=lambda item: item["value"],
    )[:10]
    top_measured_gaps = sorted(
        [
            {
                "label": record["supplier_name"] or record["part_number"] or "Unknown",
                "part_number": record["part_number"],
                "value": record["measured_capacity_gap"] or 0,
            }
            for record in records
            if (record["measured_capacity_gap"] or 0) < 0
        ],
        key=lambda item: item["value"],
    )[:10]

    cat_eval_known = sum(value for status, value in cat_evaluation_distribution.items() if status != "Unknown")
    cat_eval_red = cat_evaluation_distribution.get("Red", 0)

    priority_rows = sorted(
        [record for record in records if record["priority"] < 4],
        key=lambda record: (
            record["priority"],
            {"Red": 0, "Orange": 1, "Green": 2, "Unknown": 3}.get(str(record.get("cat_evaluation_status")), 3),
            record["measured_capacity_gap"] if record["measured_capacity_gap"] is not None else 0,
            record["capacity_gap"] if record["capacity_gap"] is not None else 0,
            0 if record["cat_status"] == "Late" else 1,
            record["supplier_name"],
        ),
    )[:80]

    def options(key: str) -> List[str]:
        values = sorted({str(record.get(key) or "").strip() for record in all_records if str(record.get(key) or "").strip()})
        return values

    return {
        "summary": {
            "total_lines": len(records),
            "suppliers_count": len({record["supplier_name"] for record in records if record["supplier_name"]}),
            "total_requested_capacity": total_requested,
            "total_contracted_capacity": total_contracted,
            "total_measured_capacity": total_measured,
            "coverage_rate": total_contracted / total_requested if total_requested else None,
            "measured_coverage_rate": total_measured / total_requested if total_requested else None,
            "missing_capacity": missing_capacity,
            "total_measured_gap": total_measured_gap,
            "missing_measured_capacity": missing_measured_capacity,
            "measured_insufficient_lines": sum(1 for record in records if (record["measured_capacity_gap"] or 0) < 0),
            "red_lines": sum(1 for record in records if record["gor"] == "Red"),
            "late_cats": sum(1 for record in records if record["cat_status"] == "Late"),
            "cat_evaluation_green": cat_evaluation_distribution.get("Green", 0),
            "cat_evaluation_orange": cat_evaluation_distribution.get("Orange", 0),
            "cat_evaluation_red": cat_eval_red,
            "cat_evaluation_unknown": cat_evaluation_distribution.get("Unknown", 0),
            "cat_evaluation_red_rate": cat_eval_red / cat_eval_known if cat_eval_known else None,
            "average_data_completeness": round(sum(record["data_completeness"] for record in records) / len(records), 1) if records else 0,
        },
        "gor_distribution": gor_distribution,
        "measured_status_distribution": measured_status_distribution,
        "capacity_comparison": {
            "requested": total_requested,
            "contracted": total_contracted,
            "measured": total_measured,
        },
        "top_risk_suppliers": _top_counts(records, "supplier_name"),
        "top_capacity_gaps": top_gaps,
        "top_measured_capacity_gaps": top_measured_gaps,
        "cat_status": cat_status,
        "cat_valuation": cat_evaluation_distribution,
        "cat_evaluation_distribution": cat_evaluation_distribution,
        "cat_evaluation_by_supplier": _stacked_risk_by(records, "supplier_name"),
        "cat_evaluation_by_sqe": _stacked_risk_by(records, "sqe"),
        "cat_evaluation_by_buyer": _stacked_risk_by(records, "buyer"),
        "risk_by_buyer": _top_counts(records, "buyer"),
        "risk_by_country": _top_counts(records, "country"),
        "priority_actions": priority_rows,
        "filters": {
            "supplier": options("supplier_name"),
            "part_number": options("part_number"),
            "country": options("country"),
            "location": options("location"),
            "buyer": options("buyer"),
            "sqe": options("sqe"),
            "gor": ["Green", "Orange", "Red", "Unknown"],
            "apqp": options("apqp_grid_project"),
            "use_cases": options("use_cases"),
            "year": options("year_of_max_need"),
            "cat_status": ["Done", "Late", "Planned", "Not planned"],
            "cat_evaluation_status": ["Green", "Orange", "Red", "Unknown"],
            "capacity_source": options("capacity_source"),
        },
    }


def _template_cell_value(row: Dict[str, Any], header: Any) -> str:
    for candidate in _template_aliases(_normalized_header(header)):
        if candidate in row:
            return _safe_excel_value(row.get(candidate))
    return ""


def _build_template_export(project_id: int) -> BytesIO:
    if not REACT_CMF_TEMPLATE_PATH.exists():
        raise HTTPException(status_code=500, detail="CMF export template not found")
    try:
        from copy import copy
        from openpyxl import load_workbook
        from openpyxl.styles import Border, PatternFill, Side
        from openpyxl.utils.cell import range_boundaries
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"openpyxl is required for CMF export: {exc}") from exc

    wb = load_workbook(REACT_CMF_TEMPLATE_PATH)
    ws = wb["CMF"] if "CMF" in wb.sheetnames else wb.active
    for table_name in list(ws.tables.keys()):
        del ws.tables[table_name]
    header_row = 1 if any(_normalized_header(ws.cell(row=2, column=col_idx).value) == "xxx" for col_idx in range(1, ws.max_column + 1)) else 3
    data_start_row = 4
    headers = [ws.cell(row=header_row, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
    rows = _project_full_rows(project_id)
    custom_columns = _project_custom_column_names(project_id)
    thin_side = Side(style="thin", color="808080")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    separator_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    clear_fill = PatternFill(fill_type=None)
    valuation_fills = {
        "G": PatternFill(fill_type="solid", fgColor="C6EFCE"),
        "O": PatternFill(fill_type="solid", fgColor="FCE4D6"),
        "R": PatternFill(fill_type="solid", fgColor="FFC7CE"),
    }
    no_fill_headers = {
        "INVESTMENT impact of Capacity Step (EURO)",
        "LEAD TIME to implement Capacity Step (WEEKS)",
        "CAPACITY CONTRACTED STEP (Parts/Week)",
    }
    valuation_headers = {GOR_REQUESTED_COLUMN, GOR_COLUMN, CAT_VALUATION_COLUMN}
    separator_columns = {
        col_idx
        for col_idx in range(1, ws.max_column + 1)
        if header_row == 1 and _normalized_header(ws.cell(row=2, column=col_idx).value).lower() == "xxx"
    }
    no_fill_columns = {
        col_idx
        for col_idx, header in enumerate(headers, 1)
        if _normalized_header(header) in no_fill_headers
    }
    for cf_range in list(ws.conditional_formatting):
        should_remove = False
        for range_part in str(cf_range.sqref).split():
            min_col, _, max_col, _ = range_boundaries(range_part)
            if any(min_col <= col_idx <= max_col for col_idx in no_fill_columns):
                should_remove = True
                break
        if should_remove:
            del ws.conditional_formatting[str(cf_range.sqref)]
    custom_placeholder_columns = [
        col_idx
        for col_idx, header in enumerate(headers, 1)
        if _is_custom_placeholder(str(header or ""))
    ]

    if header_row == 1:
        ws.row_dimensions[2].hidden = True
        ws.row_dimensions[3].hidden = True

    for col_idx, header in enumerate(headers, 1):
        normalized_header = _normalized_header(header)
        if normalized_header.startswith("Project - Part of Project"):
            ws.cell(row=header_row, column=col_idx).value = _display_template_header(normalized_header, project_id)
            headers[col_idx - 1] = ws.cell(row=header_row, column=col_idx).value

    for idx, col_idx in enumerate(custom_placeholder_columns):
        ws.cell(row=header_row, column=col_idx).value = custom_columns[idx] if idx < len(custom_columns) else ""
        headers[col_idx - 1] = ws.cell(row=header_row, column=col_idx).value

    for row_idx in range(data_start_row, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).value = None

    style_row = data_start_row
    for row_offset, data_row in enumerate(rows, data_start_row):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_offset, column=col_idx)
            source = ws.cell(row=style_row, column=col_idx)
            if row_offset != style_row and source.has_style:
                cell._style = copy(source._style)
                cell.font = copy(source.font)
                cell.fill = copy(source.fill)
                cell.border = copy(source.border)
                cell.alignment = copy(source.alignment)
                cell.number_format = source.number_format
                cell.protection = copy(source.protection)
            marker = _normalized_header(ws.cell(row=header_row + 1, column=col_idx).value) if header_row == 1 else ""
            normalized_header = _normalized_header(header)
            cell.value = "" if marker.lower() == "xxx" else _template_cell_value(data_row, header)
            cell.border = thin_border
            if col_idx in separator_columns:
                cell.fill = separator_fill
            elif normalized_header in no_fill_headers:
                cell.fill = clear_fill
            elif normalized_header in valuation_headers:
                cell.fill = valuation_fills.get(str(cell.value).strip().upper(), clear_fill)

    max_style_row = max(data_start_row + len(rows) - 1, data_start_row)
    for row_idx in range(header_row, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            normalized_header = _normalized_header(headers[col_idx - 1])
            cell.border = thin_border
            if col_idx in separator_columns:
                cell.fill = separator_fill
            elif row_idx >= data_start_row and normalized_header in no_fill_headers:
                cell.fill = clear_fill
            elif row_idx >= data_start_row and normalized_header in valuation_headers:
                cell.fill = valuation_fills.get(str(cell.value).strip().upper(), clear_fill)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _parse_delimited_file(raw: bytes, filename: str) -> Dict[str, Any]:
    text = raw.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    delimiter = "\t" if filename.lower().endswith((".tsv", ".txt")) else ","
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except Exception:
        pass
    reader = csv.DictReader(text.splitlines(), delimiter=delimiter)
    rows = [
        {str(key or "").strip(): value for key, value in row.items() if key}
        for row in reader
    ]
    columns = [str(column or "").strip() for column in (reader.fieldnames or []) if str(column or "").strip()]
    return {"columns": columns, "rows": rows[:1000], "total_rows": len(rows)}


def _parse_excel_file(raw: bytes) -> Dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"openpyxl is required to parse Excel files: {exc}") from exc

    wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = None
    for row in rows_iter:
        values = ["" if value is None else str(value).strip() for value in row]
        if any(values):
            header_row = values
            break
    if not header_row:
        return {"columns": [], "rows": [], "total_rows": 0}

    columns = []
    seen: Dict[str, int] = {}
    for idx, header in enumerate(header_row, 1):
        name = header or f"Column {idx}"
        if name in seen:
            seen[name] += 1
            name = f"{name} ({seen[name]})"
        else:
            seen[name] = 1
        columns.append(name)

    parsed_rows = []
    for row in rows_iter:
        values = ["" if value is None else str(value) for value in row]
        if not any(value.strip() for value in values):
            continue
        parsed_rows.append({columns[idx]: values[idx] if idx < len(values) else "" for idx in range(len(columns))})

    return {"columns": columns, "rows": parsed_rows[:1000], "total_rows": len(parsed_rows)}


def _parse_import_file(payload: ImportFileParse) -> Dict[str, Any]:
    try:
        raw = base64.b64decode(payload.content_base64)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid file content: {exc}") from exc
    filename = payload.filename.lower()
    if filename.endswith((".xlsx", ".xlsm")):
        return _parse_excel_file(raw)
    if filename.endswith((".csv", ".tsv", ".txt")):
        return _parse_delimited_file(raw, filename)
    raise HTTPException(status_code=400, detail="Unsupported file type. Use Excel (.xlsx/.xlsm), CSV, TSV, or TXT.")


@app.on_event("startup")
def startup() -> None:
    db_sqlite.init_db()
    _ensure_users_table()
    _ensure_react_schema_migrations()


@app.get("/api/health")
def health() -> Dict[str, str]:
    return {"status": "ok", "database": str(db_sqlite.DB_PATH)}


@app.get("/api/schema")
def schema() -> Dict[str, Any]:
    return {
        "columns": [
            {
                "name": col["name"],
                "section": col["section"],
                "roles": col["roles"],
                "is_auto": col["is_auto"],
            }
            for col in STANDARD_COLUMNS_WITH_ROLES
        ]
    }


@app.get("/api/projects")
def list_projects() -> Dict[str, Any]:
    return {"projects": [_project_payload(project) for project in project_repo.get_all_projects()]}


@app.post("/api/projects")
def create_project(payload: ProjectCreate) -> Dict[str, Any]:
    project_name = payload.project.strip()
    part_name = payload.part_of_project.strip()
    if not project_name or not part_name:
        raise HTTPException(status_code=400, detail="project and part_of_project are required")

    code = _project_code(project_name, part_name)
    if project_repo.get_project_by_code(code):
        raise HTTPException(status_code=409, detail="This Projet / Part of Project combination already exists")

    project = project_repo.create_project(
        code=code,
        name=part_name,
        capacity_manager_name=payload.capacity_manager_name.strip() or "capacity_manager",
        description=payload.description,
        buyer_assigned_name=payload.buyer_assigned_name,
        sqd_assigned_name=payload.sqd_assigned_name,
        cmf_status=payload.cmf_status,
        created_by=payload.created_by,
    )
    project_repo.assign_user_to_project(payload.capacity_manager_name, project.id, "CAPACITY_MANAGER", assigned_by=payload.created_by)
    if payload.buyer_assigned_name:
        project_repo.assign_user_to_project(payload.buyer_assigned_name, project.id, "BUYER", assigned_by=payload.created_by)
    if payload.sqd_assigned_name:
        project_repo.assign_user_to_project(payload.sqd_assigned_name, project.id, "SQD", assigned_by=payload.created_by)
    for custom_column in payload.custom_columns:
        _add_custom_column_to_project_section(project.id, custom_column.column_name, custom_column.owner_role, custom_column.section)
    return {"project": _project_payload(project)}


@app.get("/api/projects/{project_id}/columns")
def project_columns(project_id: int) -> Dict[str, Any]:
    if not project_repo.get_project_by_id(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    columns = column_repo.get_columns(project_id)
    return {
        "columns": [
            {
                **column,
                "section": column.get("custom_section") or get_column_section(column["column_name"]),
                "is_custom": get_column_section(column["column_name"]) == "CUSTOMIZED COLUMNS",
                "roles": column_repo.get_roles_for_column(project_id, column["column_name"]),
            }
            for column in columns
        ]
    }


@app.post("/api/projects/{project_id}/custom-columns")
def add_project_custom_column(project_id: int, payload: CustomColumnPayload) -> Dict[str, Any]:
    project = project_repo.get_project_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    _ensure_project_write_access(project, payload.actor_email, "CAPACITY_MANAGER")
    return {"column": _add_custom_column_to_project_section(project_id, payload.column_name, payload.owner_role, payload.section)}


@app.get("/api/projects/{project_id}/editable-columns")
def editable_columns(project_id: int, role: str, section: Optional[str] = None) -> Dict[str, Any]:
    if not project_repo.get_project_by_id(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    columns = column_repo.get_columns_for_role(project_id, role)
    if section:
        columns = [column for column in columns if _column_section_for_project(project_id, column) == section]
    return {"columns": [{"name": column, "section": _column_section_for_project(project_id, column)} for column in columns]}


@app.get("/api/projects/{project_id}/records")
def project_records(project_id: int) -> Dict[str, Any]:
    if not project_repo.get_project_by_id(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    return {"records": [_record_payload(record) for record in record_repo.get_records_for_project(project_id)]}


@app.get("/api/projects/{project_id}/full-data")
def project_full_data(project_id: int) -> Dict[str, Any]:
    if not project_repo.get_project_by_id(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    columns = _template_columns()
    rows = _project_full_rows(project_id)
    return {
        "columns": columns,
        "default_visible": [column for column in REACT_DEFAULT_VIEW_COLS if column in columns],
        "records": _display_rows(rows, columns),
    }


@app.get("/api/projects/{project_id}/dashboard")
def project_dashboard(
    project_id: int,
    supplier: Optional[str] = None,
    part_number: Optional[str] = None,
    country: Optional[str] = None,
    location: Optional[str] = None,
    buyer: Optional[str] = None,
    sqe: Optional[str] = None,
    gor: Optional[str] = None,
    apqp: Optional[str] = None,
    use_cases: Optional[str] = None,
    year: Optional[str] = None,
    cat_status: Optional[str] = None,
    cat_evaluation_status: Optional[str] = None,
    capacity_source: Optional[str] = None,
) -> Dict[str, Any]:
    if not project_repo.get_project_by_id(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    filters = {
        "supplier": supplier,
        "part_number": part_number,
        "country": country,
        "location": location,
        "buyer": buyer,
        "sqe": sqe,
        "gor": gor,
        "apqp": apqp,
        "use_cases": use_cases,
        "year": year,
        "cat_status": cat_status,
        "cat_evaluation_status": cat_evaluation_status,
        "capacity_source": capacity_source,
    }
    return _dashboard_payload(project_id, filters)


@app.get("/api/cmf/dashboard/summary")
def dashboard_summary(project_id: int, cat_evaluation_status: Optional[str] = None) -> Dict[str, Any]:
    if not project_repo.get_project_by_id(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    return {"summary": _dashboard_payload(project_id, {"cat_evaluation_status": cat_evaluation_status})["summary"]}


@app.get("/api/cmf/dashboard/gor-distribution")
def dashboard_gor_distribution(project_id: int) -> Dict[str, Any]:
    if not project_repo.get_project_by_id(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    return {"gor_distribution": _dashboard_payload(project_id, {})["gor_distribution"]}


@app.get("/api/cmf/dashboard/top-risk-suppliers")
def dashboard_top_risk_suppliers(project_id: int) -> Dict[str, Any]:
    if not project_repo.get_project_by_id(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    return {"top_risk_suppliers": _dashboard_payload(project_id, {})["top_risk_suppliers"]}


@app.get("/api/cmf/dashboard/capacity-gaps")
def dashboard_capacity_gaps(project_id: int) -> Dict[str, Any]:
    if not project_repo.get_project_by_id(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    return {"top_capacity_gaps": _dashboard_payload(project_id, {})["top_capacity_gaps"]}


@app.get("/api/cmf/dashboard/cat-status")
def dashboard_cat_status(project_id: int) -> Dict[str, Any]:
    if not project_repo.get_project_by_id(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    return {"cat_status": _dashboard_payload(project_id, {})["cat_status"]}


@app.get("/api/cmf/dashboard/action-plan")
def dashboard_action_plan(project_id: int, cat_evaluation_status: Optional[str] = None) -> Dict[str, Any]:
    if not project_repo.get_project_by_id(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    return {"priority_actions": _dashboard_payload(project_id, {"cat_evaluation_status": cat_evaluation_status})["priority_actions"]}


@app.get("/api/cmf/dashboard/filters")
def dashboard_filters(project_id: int) -> Dict[str, Any]:
    if not project_repo.get_project_by_id(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    return {"filters": _dashboard_payload(project_id, {})["filters"]}


@app.get("/api/cmf/dashboard/measured-capacity-gaps")
def dashboard_measured_capacity_gaps(project_id: int, cat_evaluation_status: Optional[str] = None) -> Dict[str, Any]:
    if not project_repo.get_project_by_id(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    return {"top_measured_capacity_gaps": _dashboard_payload(project_id, {"cat_evaluation_status": cat_evaluation_status})["top_measured_capacity_gaps"]}


@app.get("/api/cmf/dashboard/cat-evaluation-distribution")
def dashboard_cat_evaluation_distribution(project_id: int, cat_evaluation_status: Optional[str] = None) -> Dict[str, Any]:
    if not project_repo.get_project_by_id(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    return {"cat_evaluation_distribution": _dashboard_payload(project_id, {"cat_evaluation_status": cat_evaluation_status})["cat_evaluation_distribution"]}


@app.patch("/api/projects/{project_id}/records/{record_id}")
def update_project_record(project_id: int, record_id: int, payload: ProjectRecordDirectSave) -> Dict[str, Any]:
    project = project_repo.get_project_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    _ensure_project_write_access(project, payload.actor_email)
    existing = record_repo.get_record_by_id(record_id)
    if not existing or existing.project_id != project_id:
        raise HTTPException(status_code=404, detail="Record not found")

    data = dict(payload.values)
    if payload.part_number is not None:
        data["part_number"] = payload.part_number
    if payload.apqp_grid is not None:
        data["apqp_grid"] = payload.apqp_grid
    updated = record_repo.update_record(record_id, data, payload.updated_by)
    _write_audit_log(
        "PROJECT_UPDATE_RECORD",
        "cmf_record",
        record_id,
        payload.updated_by,
        str(existing.to_dict()),
        str(data),
        project_id,
    )
    return {"record": _record_payload(updated)}


@app.delete("/api/projects/{project_id}/records/{record_id}")
def delete_project_record(project_id: int, record_id: int, actor_email: Optional[str] = None, updated_by: str = "react-ui") -> Dict[str, Any]:
    project = project_repo.get_project_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    _ensure_project_write_access(project, actor_email)
    existing = record_repo.get_record_by_id(record_id)
    if not existing or existing.project_id != project_id:
        raise HTTPException(status_code=404, detail="Record not found")
    if not record_repo.delete_record(record_id):
        raise HTTPException(status_code=404, detail="Record not found")
    _write_audit_log(
        "PROJECT_DELETE_RECORD",
        "cmf_record",
        record_id,
        updated_by,
        str(existing.to_dict()),
        None,
        project_id,
    )
    return {"deleted": True}


@app.get("/api/projects/{project_id}/action-plans")
def list_action_plans(project_id: int) -> Dict[str, Any]:
    if not project_repo.get_project_by_id(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    return {"action_plans": _action_plan_rows(project_id)}


@app.post("/api/projects/{project_id}/action-plans")
def create_action_plan(project_id: int, payload: ActionPlanCreate) -> Dict[str, Any]:
    project = project_repo.get_project_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    _ensure_project_write_access(project, payload.actor_email, "SQD")
    record = record_repo.get_record_by_id(payload.record_id)
    if not record or record.project_id != project_id:
        raise HTTPException(status_code=404, detail="Record not found")

    problem = payload.problem_type.strip()
    if problem == "CAT Requested vs Measured":
        target_column = MEASURED_COLUMN
    elif problem == "GOR Requested vs Contracted":
        target_column = CONTRACTED_COLUMN
    else:
        raise HTTPException(status_code=400, detail="Invalid action plan problem type")

    action_text = payload.action_text.strip()
    new_value = str(payload.new_value).strip()
    if not action_text:
        raise HTTPException(status_code=400, detail="Action text is required")
    if not new_value:
        raise HTTPException(status_code=400, detail="New value is required")

    old_value = payload.old_value
    if old_value in (None, ""):
        old_value = str(record.to_dict().get(target_column) or "")

    updated = record_repo.update_record(payload.record_id, {target_column: new_value}, payload.created_by)
    conn = db_sqlite.get_connection()
    try:
        with conn:
            cur = conn.execute(
                """
                INSERT INTO action_plans (
                    project_id, record_id, part_number, problem_type, target_column,
                    action_text, old_value, new_value, created_by
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_id,
                    payload.record_id,
                    record.part_number,
                    problem,
                    target_column,
                    action_text,
                    old_value,
                    new_value,
                    payload.created_by,
                ),
            )
            action_plan_id = cur.lastrowid
    finally:
        db_sqlite.close_connection(conn)

    _write_audit_log(
        "SQD_CREATE_ACTION_PLAN",
        "action_plan",
        action_plan_id,
        payload.created_by,
        f"{target_column}: {old_value}",
        f"{target_column}: {new_value}; action: {action_text}",
        project_id,
    )
    return {
        "action_plan": next((plan for plan in _action_plan_rows(project_id) if plan["id"] == action_plan_id), None),
        "record": _record_payload(updated),
    }


@app.get("/api/projects/{project_id}/cmf-export")
def export_project_cmf(project_id: int):
    if not project_repo.get_project_by_id(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    output = _build_template_export(project_id)
    filename = f"CMF_export_project_{project_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/cross-project")
def cross_project_view() -> Dict[str, Any]:
    projects = record_repo.get_all_projects_for_cross_view()
    cross_data = record_repo.get_cross_project_part_number_view()
    rows = []
    for entry in cross_data:
        project_count = int(entry.get("project_count") or sum(1 for key, value in entry.items() if key.startswith("proj_") and value))
        row = {
            "APQP": entry.get("apqp") or "",
            "Part Name": entry.get("part_name") or "",
            "Part Number": entry.get("part_number") or "",
            "CarryOver - Adapted": "Adapted" if project_count == 1 else "CarryOver",
        }
        for project in projects:
            row[_project_label(project)] = bool(entry.get(f"proj_{project['id']}", False))
        rows.append(row)
    return {"projects": [_project_label(project) for project in projects], "records": rows}


@app.post("/api/import/parse-file")
def parse_import_file(payload: ImportFileParse) -> Dict[str, Any]:
    return _parse_import_file(payload)


@app.post("/api/projects/{project_id}/records")
def upsert_record(project_id: int, payload: RecordUpsert) -> Dict[str, Any]:
    project = project_repo.get_project_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    _ensure_project_write_access(project, payload.actor_email)
    values = dict(payload.values)
    values["part_number"] = payload.part_number
    if payload.apqp_grid:
        values["apqp_grid"] = payload.apqp_grid
    record_id = record_repo.ensure_record(
        project_id=project_id,
        part_number=payload.part_number,
        values=values,
        updated_by=payload.updated_by,
        apqp=payload.apqp_grid,
    )
    record = record_repo.get_record_by_id(record_id)
    return {"record": _record_payload(record)}


@app.post("/api/projects/{project_id}/role-record")
def save_role_record(project_id: int, payload: RoleRecordSave) -> Dict[str, Any]:
    project = project_repo.get_project_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    _ensure_project_write_access(project, payload.actor_email, payload.role.upper())

    editable = set(column_repo.get_columns_for_role(project_id, payload.role))
    if payload.section:
        editable = {column for column in editable if _column_section_for_project(project_id, column) == payload.section}

    values = {
        key: value
        for key, value in payload.values.items()
        if key in editable and key not in {"PART NUMBER"}
    }
    part_number = payload.part_number.strip()
    if not part_number:
        raise HTTPException(status_code=400, detail="PART NUMBER is required")

    existing_id = record_repo.find_record_by_part_number(project_id, part_number)
    if existing_id:
        if payload.apqp_grid and "APQP GRID" in editable:
            values["apqp_grid"] = payload.apqp_grid
        record_repo.update_record(existing_id, {"part_number": part_number, **values}, payload.updated_by)
        record = record_repo.get_record_by_id(existing_id)
        return {"record": _record_payload(record)}

    if not payload.create_if_missing:
        raise HTTPException(status_code=404, detail="Part Number not found")

    create_values = {"part_number": part_number, **values}
    if payload.apqp_grid and "APQP GRID" in editable:
        create_values["apqp_grid"] = payload.apqp_grid

    record_id = record_repo.ensure_record(
        project_id=project_id,
        part_number=part_number,
        values=create_values,
        updated_by=payload.updated_by,
        apqp=payload.apqp_grid,
    )
    record = record_repo.get_record_by_id(record_id)
    return {"record": _record_payload(record)}


@app.patch("/api/records/{record_id}/values")
def update_record_values(record_id: int, payload: ValueUpdate) -> Dict[str, Any]:
    record = record_repo.get_record_by_id(record_id)
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    record_repo.update_values(record_id, payload.values, payload.updated_by)
    updated = record_repo.get_record_by_id(record_id)
    return {"record": _record_payload(updated)}


@app.patch("/api/projects/{project_id}")
def update_project(project_id: int, payload: ProjectUpdate) -> Dict[str, Any]:
    current_project = project_repo.get_project_by_id(project_id)
    if not current_project:
        raise HTTPException(status_code=404, detail="Project not found")
    _ensure_project_write_access(current_project, payload.actor_email, "CAPACITY_MANAGER")
    data = {key: value for key, value in payload.model_dump().items() if value is not None and key != "actor_email"}
    project = project_repo.update_project(project_id, **data)
    return {"project": _project_payload(project)}


@app.delete("/api/projects/{project_id}")
def delete_project(project_id: int) -> Dict[str, Any]:
    project = project_repo.get_project_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if not project_repo.delete_project(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    _write_audit_log(
        "ADMIN_DELETE_PROJECT",
        "project",
        project_id,
        "admin",
        str(_project_payload(project)),
        None,
        project_id,
    )
    return {"deleted": True}


@app.delete("/api/records/{record_id}")
def delete_record(record_id: int) -> Dict[str, Any]:
    if not record_repo.delete_record(record_id):
        raise HTTPException(status_code=404, detail="Record not found")
    return {"deleted": True}


@app.get("/api/admin/audit-logs")
def admin_audit_logs(limit: int = 250) -> Dict[str, Any]:
    return {"logs": _audit_rows(limit)}


@app.post("/api/admin/projects/{project_id}/records")
def admin_upsert_record(project_id: int, payload: AdminRecordDirectSave) -> Dict[str, Any]:
    if not project_repo.get_project_by_id(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    values = dict(payload.values)
    values["part_number"] = payload.part_number
    if payload.apqp_grid is not None:
        values["apqp_grid"] = payload.apqp_grid
    record_id = record_repo.ensure_record(
        project_id=project_id,
        part_number=payload.part_number,
        values=values,
        updated_by=payload.updated_by,
        apqp=payload.apqp_grid,
    )
    record = record_repo.get_record_by_id(record_id)
    _write_audit_log(
        "ADMIN_UPSERT_RECORD",
        "cmf_record",
        record_id,
        payload.updated_by,
        None,
        str(values),
        project_id,
    )
    return {"record": _record_payload(record)}


@app.patch("/api/admin/records/{record_id}")
def admin_update_record(record_id: int, payload: AdminRecordDirectSave) -> Dict[str, Any]:
    existing = record_repo.get_record_by_id(record_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Record not found")
    data = {"part_number": payload.part_number, "apqp_grid": payload.apqp_grid, **payload.values}
    updated = record_repo.update_record(record_id, data, payload.updated_by)
    _write_audit_log(
        "ADMIN_UPDATE_RECORD",
        "cmf_record",
        record_id,
        payload.updated_by,
        str(existing.to_dict()),
        str(data),
        existing.project_id,
    )
    return {"record": _record_payload(updated)}


@app.delete("/api/admin/records/{record_id}")
def admin_delete_record(record_id: int, updated_by: str = "admin") -> Dict[str, Any]:
    existing = record_repo.get_record_by_id(record_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Record not found")
    project_id = existing.project_id
    if not record_repo.delete_record(record_id):
        raise HTTPException(status_code=404, detail="Record not found")
    _write_audit_log(
        "ADMIN_DELETE_RECORD",
        "cmf_record",
        record_id,
        updated_by,
        str(existing.to_dict()),
        None,
        project_id,
    )
    return {"deleted": True}


@app.delete("/api/admin/projects/{project_id}/records")
def admin_reset_project_records(project_id: int, updated_by: str = "admin") -> Dict[str, Any]:
    if not project_repo.get_project_by_id(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    conn = db_sqlite.get_connection()
    try:
        with conn:
            before = conn.execute("SELECT COUNT(*) AS count FROM cmf_records WHERE project_id = ?", (project_id,)).fetchone()["count"]
            conn.execute("DELETE FROM cmf_records WHERE project_id = ?", (project_id,))
    finally:
        db_sqlite.close_connection(conn)
    _write_audit_log(
        "ADMIN_RESET_PROJECT_RECORDS",
        "project",
        project_id,
        updated_by,
        f"{before} records",
        "0 records",
        project_id,
    )
    return {"deleted_records": before}


@app.get("/api/users")
def list_users() -> Dict[str, Any]:
    _ensure_users_table()
    return {"users": [_user_payload(user) for user in user_repo.list_all_users()]}


@app.post("/api/auth/login")
def login(payload: LoginPayload) -> Dict[str, Any]:
    _ensure_users_table()
    user = user_repo.get_user_by_email(payload.email.strip().lower())
    if not user or not _verify_password(payload.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    return {"user": _user_payload(user)}


@app.post("/api/users")
def create_user(payload: UserCreate) -> Dict[str, Any]:
    _ensure_users_table()
    if user_repo.user_exists(payload.email):
        raise HTTPException(status_code=409, detail="User already exists")
    if len(payload.password) < 6:
        raise HTTPException(status_code=400, detail="Password must contain at least 6 characters")
    user = user_repo.create_user(
        email=payload.email.strip().lower(),
        password_hash=_hash_password(payload.password),
        full_name=payload.full_name,
    )
    if not user:
        raise HTTPException(status_code=400, detail="Unable to create user")
    user_repo.update_user_role(user.id, payload.role)
    updated = user_repo.get_user_by_id(user.id)
    return {"user": _user_payload(updated)}


@app.patch("/api/users/{user_id}/role")
def update_user_role(user_id: int, payload: UserRoleUpdate) -> Dict[str, Any]:
    _ensure_users_table()
    if not user_repo.update_user_role(user_id, payload.role):
        raise HTTPException(status_code=404, detail="User not found")
    user = user_repo.get_user_by_id(user_id)
    return {"user": _user_payload(user)}


@app.patch("/api/users/{user_id}/password")
def update_user_password(user_id: int, payload: UserPasswordUpdate) -> Dict[str, Any]:
    _ensure_users_table()
    if len(payload.password) < 6:
        raise HTTPException(status_code=400, detail="Password must contain at least 6 characters")
    user = user_repo.update_user(user_id, password_hash=_hash_password(payload.password))
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    _write_audit_log("ADMIN_UPDATE_USER_PASSWORD", "app_user", user_id, "admin", None, "password updated", None)
    return {"user": _user_payload(user)}


@app.delete("/api/users/{user_id}")
def delete_user(user_id: int) -> Dict[str, Any]:
    _ensure_users_table()
    user = user_repo.get_user_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if not user_repo.delete_user(user_id):
        raise HTTPException(status_code=404, detail="User not found")
    _write_audit_log("ADMIN_DELETE_USER", "app_user", user_id, "admin", str(_user_payload(user)), None, None)
    return {"deleted": True}
